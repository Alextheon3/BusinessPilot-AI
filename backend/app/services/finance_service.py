from typing import Optional, List, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, extract, and_
from datetime import datetime, timedelta
from app.models.finance import Expense
from app.schemas.finance import ExpenseCreate, ExpenseUpdate
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
import io


class FinanceService:
    def __init__(self, db: Session):
        self.db = db

    def create_expense(self, expense_create: ExpenseCreate) -> Expense:
        db_expense = Expense(
            amount=expense_create.amount,
            category=expense_create.category,
            description=expense_create.description,
            date=expense_create.date,
            receipt_url=expense_create.receipt_url,
            is_recurring=expense_create.is_recurring
        )
        self.db.add(db_expense)
        self.db.commit()
        self.db.refresh(db_expense)
        return db_expense

    def get_expense(self, expense_id: int) -> Optional[Expense]:
        return self.db.query(Expense).filter(Expense.id == expense_id).first()

    def get_expenses(self, skip: int = 0, limit: int = 100,
                    start_date: Optional[datetime] = None,
                    end_date: Optional[datetime] = None,
                    category: Optional[str] = None) -> List[Expense]:
        query = self.db.query(Expense)
        
        if start_date:
            query = query.filter(Expense.date >= start_date.date())
        if end_date:
            query = query.filter(Expense.date <= end_date.date())
        if category:
            query = query.filter(Expense.category == category)
        
        return query.order_by(desc(Expense.date)).offset(skip).limit(limit).all()

    def update_expense(self, expense_id: int, expense_update: ExpenseUpdate) -> Optional[Expense]:
        db_expense = self.get_expense(expense_id)
        if not db_expense:
            return None
        
        update_data = expense_update.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_expense, field, value)
        
        self.db.commit()
        self.db.refresh(db_expense)
        return db_expense

    def delete_expense(self, expense_id: int) -> bool:
        db_expense = self.get_expense(expense_id)
        if not db_expense:
            return False
        
        self.db.delete(db_expense)
        self.db.commit()
        return True

    def get_expense_categories(self) -> List[str]:
        """Get all expense categories"""
        categories = self.db.query(Expense.category).distinct().all()
        return [cat[0] for cat in categories if cat[0]]

    def get_financial_summary(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Get comprehensive financial summary"""
        
        # Get revenue from sales
        revenue = self._get_revenue(start_date, end_date)
        
        # Get expenses
        expenses = self._get_expenses_summary(start_date, end_date)
        
        # Calculate profit/loss
        total_revenue = revenue.get("total_revenue", 0)
        total_expenses = expenses.get("total_expenses", 0)
        profit_loss = total_revenue - total_expenses
        
        # Get expense breakdown by category
        expense_breakdown = self._get_expense_breakdown(start_date, end_date)
        
        # Get monthly trends
        monthly_trends = self._get_monthly_trends(start_date, end_date)
        
        return {
            "period": {
                "start_date": start_date.strftime('%Y-%m-%d'),
                "end_date": end_date.strftime('%Y-%m-%d')
            },
            "summary": {
                "total_revenue": total_revenue,
                "total_expenses": total_expenses,
                "profit_loss": profit_loss,
                "profit_margin": (profit_loss / total_revenue * 100) if total_revenue > 0 else 0,
                "expense_ratio": (total_expenses / total_revenue * 100) if total_revenue > 0 else 0
            },
            "revenue": revenue,
            "expenses": expenses,
            "expense_breakdown": expense_breakdown,
            "monthly_trends": monthly_trends
        }

    def _get_revenue(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Get revenue data from sales"""
        
        try:
            from app.models.sales import Sale
            
            sales_query = self.db.query(Sale).filter(
                Sale.created_at >= start_date,
                Sale.created_at <= end_date
            )
            
            sales = sales_query.all()
            
            total_revenue = sum(sale.total_amount for sale in sales)
            total_sales = len(sales)
            average_sale = total_revenue / total_sales if total_sales > 0 else 0
            
            # Get daily revenue
            daily_revenue = self.db.query(
                func.date(Sale.created_at).label('date'),
                func.sum(Sale.total_amount).label('revenue')
            ).filter(
                Sale.created_at >= start_date,
                Sale.created_at <= end_date
            ).group_by(func.date(Sale.created_at)).order_by('date').all()
            
            return {
                "total_revenue": total_revenue,
                "total_sales": total_sales,
                "average_sale": average_sale,
                "daily_revenue": [
                    {
                        "date": day.date.isoformat(),
                        "revenue": day.revenue
                    }
                    for day in daily_revenue
                ]
            }
        except Exception as e:
            print(f"Error getting revenue data: {e}")
            return {"total_revenue": 0, "total_sales": 0, "average_sale": 0, "daily_revenue": []}

    def _get_expenses_summary(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Get expenses summary"""
        
        expenses_query = self.db.query(Expense).filter(
            Expense.date >= start_date.date(),
            Expense.date <= end_date.date()
        )
        
        expenses = expenses_query.all()
        
        total_expenses = sum(expense.amount for expense in expenses)
        total_transactions = len(expenses)
        average_expense = total_expenses / total_transactions if total_transactions > 0 else 0
        
        # Get daily expenses
        daily_expenses = self.db.query(
            Expense.date,
            func.sum(Expense.amount).label('amount')
        ).filter(
            Expense.date >= start_date.date(),
            Expense.date <= end_date.date()
        ).group_by(Expense.date).order_by(Expense.date).all()
        
        return {
            "total_expenses": total_expenses,
            "total_transactions": total_transactions,
            "average_expense": average_expense,
            "daily_expenses": [
                {
                    "date": day.date.isoformat(),
                    "amount": day.amount
                }
                for day in daily_expenses
            ]
        }

    def _get_expense_breakdown(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """Get expense breakdown by category"""
        
        breakdown = self.db.query(
            Expense.category,
            func.sum(Expense.amount).label('total_amount'),
            func.count(Expense.id).label('transaction_count')
        ).filter(
            Expense.date >= start_date.date(),
            Expense.date <= end_date.date()
        ).group_by(Expense.category).order_by(desc('total_amount')).all()
        
        total_expenses = sum(item.total_amount for item in breakdown)
        
        return [
            {
                "category": item.category,
                "total_amount": item.total_amount,
                "transaction_count": item.transaction_count,
                "percentage": (item.total_amount / total_expenses * 100) if total_expenses > 0 else 0
            }
            for item in breakdown
        ]

    def _get_monthly_trends(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """Get monthly financial trends"""
        
        # Get monthly revenue
        try:
            from app.models.sales import Sale
            
            monthly_revenue = self.db.query(
                extract('year', Sale.created_at).label('year'),
                extract('month', Sale.created_at).label('month'),
                func.sum(Sale.total_amount).label('revenue')
            ).filter(
                Sale.created_at >= start_date,
                Sale.created_at <= end_date
            ).group_by(
                extract('year', Sale.created_at),
                extract('month', Sale.created_at)
            ).order_by('year', 'month').all()
            
        except Exception:
            monthly_revenue = []
        
        # Get monthly expenses
        monthly_expenses = self.db.query(
            extract('year', Expense.date).label('year'),
            extract('month', Expense.date).label('month'),
            func.sum(Expense.amount).label('expenses')
        ).filter(
            Expense.date >= start_date.date(),
            Expense.date <= end_date.date()
        ).group_by(
            extract('year', Expense.date),
            extract('month', Expense.date)
        ).order_by('year', 'month').all()
        
        # Combine revenue and expenses by month
        trends = {}
        
        for rev in monthly_revenue:
            key = f"{int(rev.year)}-{int(rev.month):02d}"
            trends[key] = {"revenue": rev.revenue, "expenses": 0}
        
        for exp in monthly_expenses:
            key = f"{int(exp.year)}-{int(exp.month):02d}"
            if key in trends:
                trends[key]["expenses"] = exp.expenses
            else:
                trends[key] = {"revenue": 0, "expenses": exp.expenses}
        
        # Convert to list and calculate profit
        result = []
        for period, data in sorted(trends.items()):
            profit = data["revenue"] - data["expenses"]
            result.append({
                "period": period,
                "revenue": data["revenue"],
                "expenses": data["expenses"],
                "profit": profit
            })
        
        return result

    def get_cash_flow_analysis(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Get cash flow analysis"""
        
        # Get daily cash flow
        daily_cash_flow = []
        current_date = start_date
        
        while current_date <= end_date:
            # Get revenue for the day
            try:
                from app.models.sales import Sale
                
                daily_revenue = self.db.query(
                    func.sum(Sale.total_amount)
                ).filter(
                    func.date(Sale.created_at) == current_date.date()
                ).scalar() or 0
                
            except Exception:
                daily_revenue = 0
            
            # Get expenses for the day
            daily_expenses = self.db.query(
                func.sum(Expense.amount)
            ).filter(
                Expense.date == current_date.date()
            ).scalar() or 0
            
            net_cash_flow = daily_revenue - daily_expenses
            
            daily_cash_flow.append({
                "date": current_date.strftime('%Y-%m-%d'),
                "revenue": daily_revenue,
                "expenses": daily_expenses,
                "net_cash_flow": net_cash_flow
            })
            
            current_date += timedelta(days=1)
        
        # Calculate cumulative cash flow
        cumulative_cash_flow = 0
        for day in daily_cash_flow:
            cumulative_cash_flow += day["net_cash_flow"]
            day["cumulative_cash_flow"] = cumulative_cash_flow
        
        # Calculate cash flow metrics
        total_inflow = sum(day["revenue"] for day in daily_cash_flow)
        total_outflow = sum(day["expenses"] for day in daily_cash_flow)
        net_cash_flow = total_inflow - total_outflow
        
        return {
            "summary": {
                "total_inflow": total_inflow,
                "total_outflow": total_outflow,
                "net_cash_flow": net_cash_flow,
                "cash_flow_ratio": (total_inflow / total_outflow) if total_outflow > 0 else 0
            },
            "daily_cash_flow": daily_cash_flow
        }

    def get_budget_analysis(self, budget_data: Dict[str, float], 
                          start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Analyze actual expenses vs budget"""
        
        actual_expenses = self._get_expense_breakdown(start_date, end_date)
        
        budget_analysis = []
        total_budget = sum(budget_data.values())
        total_actual = sum(item["total_amount"] for item in actual_expenses)
        
        for category, budgeted_amount in budget_data.items():
            actual_item = next((item for item in actual_expenses if item["category"] == category), None)
            actual_amount = actual_item["total_amount"] if actual_item else 0
            
            variance = actual_amount - budgeted_amount
            variance_percentage = (variance / budgeted_amount * 100) if budgeted_amount > 0 else 0
            
            budget_analysis.append({
                "category": category,
                "budgeted_amount": budgeted_amount,
                "actual_amount": actual_amount,
                "variance": variance,
                "variance_percentage": variance_percentage,
                "status": "over" if variance > 0 else "under" if variance < 0 else "on_track"
            })
        
        return {
            "summary": {
                "total_budget": total_budget,
                "total_actual": total_actual,
                "total_variance": total_actual - total_budget,
                "budget_utilization": (total_actual / total_budget * 100) if total_budget > 0 else 0
            },
            "category_analysis": budget_analysis
        }

    def export_financial_report(self, format: str, start_date: datetime, 
                              end_date: datetime) -> bytes:
        """Export financial report in PDF or Excel format"""
        
        financial_data = self.get_financial_summary(start_date, end_date)
        
        if format.lower() == "pdf":
            return self._export_pdf_report(financial_data)
        elif format.lower() == "excel":
            return self._export_excel_report(financial_data)
        else:
            raise ValueError("Unsupported format. Use 'pdf' or 'excel'")

    def _export_pdf_report(self, data: Dict[str, Any]) -> bytes:
        """Export financial report as PDF"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph("Financial Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Period
        period_text = f"Period: {data['period']['start_date']} to {data['period']['end_date']}"
        period = Paragraph(period_text, styles['Normal'])
        story.append(period)
        story.append(Spacer(1, 20))
        
        # Summary Table
        summary_data = [
            ['Metric', 'Amount'],
            ['Total Revenue', f"${data['summary']['total_revenue']:.2f}"],
            ['Total Expenses', f"${data['summary']['total_expenses']:.2f}"],
            ['Profit/Loss', f"${data['summary']['profit_loss']:.2f}"],
            ['Profit Margin', f"{data['summary']['profit_margin']:.2f}%"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Expense Breakdown
        if data['expense_breakdown']:
            expense_title = Paragraph("Expense Breakdown by Category", styles['Heading2'])
            story.append(expense_title)
            story.append(Spacer(1, 10))
            
            expense_data = [['Category', 'Amount', 'Percentage']]
            for item in data['expense_breakdown']:
                expense_data.append([
                    item['category'],
                    f"${item['total_amount']:.2f}",
                    f"{item['percentage']:.1f}%"
                ])
            
            expense_table = Table(expense_data)
            expense_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(expense_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    def _export_excel_report(self, data: Dict[str, Any]) -> bytes:
        """Export financial report as Excel"""
        
        buffer = io.BytesIO()
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Financial Summary"
        
        # Headers
        ws_summary['A1'] = "Financial Report"
        ws_summary['A2'] = f"Period: {data['period']['start_date']} to {data['period']['end_date']}"
        
        # Summary data
        ws_summary['A4'] = "Metric"
        ws_summary['B4'] = "Amount"
        ws_summary['A5'] = "Total Revenue"
        ws_summary['B5'] = data['summary']['total_revenue']
        ws_summary['A6'] = "Total Expenses"
        ws_summary['B6'] = data['summary']['total_expenses']
        ws_summary['A7'] = "Profit/Loss"
        ws_summary['B7'] = data['summary']['profit_loss']
        ws_summary['A8'] = "Profit Margin"
        ws_summary['B8'] = f"{data['summary']['profit_margin']:.2f}%"
        
        # Expense Breakdown Sheet
        if data['expense_breakdown']:
            ws_expenses = wb.create_sheet("Expense Breakdown")
            ws_expenses['A1'] = "Category"
            ws_expenses['B1'] = "Amount"
            ws_expenses['C1'] = "Percentage"
            
            for idx, item in enumerate(data['expense_breakdown'], 2):
                ws_expenses[f'A{idx}'] = item['category']
                ws_expenses[f'B{idx}'] = item['total_amount']
                ws_expenses[f'C{idx}'] = f"{item['percentage']:.1f}%"
        
        # Monthly Trends Sheet
        if data['monthly_trends']:
            ws_trends = wb.create_sheet("Monthly Trends")
            ws_trends['A1'] = "Period"
            ws_trends['B1'] = "Revenue"
            ws_trends['C1'] = "Expenses"
            ws_trends['D1'] = "Profit"
            
            for idx, item in enumerate(data['monthly_trends'], 2):
                ws_trends[f'A{idx}'] = item['period']
                ws_trends[f'B{idx}'] = item['revenue']
                ws_trends[f'C{idx}'] = item['expenses']
                ws_trends[f'D{idx}'] = item['profit']
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def get_tax_report(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Generate tax report for the specified period"""
        
        # Get revenue data
        revenue_data = self._get_revenue(start_date, end_date)
        
        # Get deductible expenses
        deductible_expenses = self.db.query(Expense).filter(
            Expense.date >= start_date.date(),
            Expense.date <= end_date.date(),
            Expense.category.in_([
                'Office Supplies', 'Marketing', 'Professional Services',
                'Equipment', 'Travel', 'Utilities', 'Insurance', 'Other'
            ])
        ).all()
        
        total_deductible = sum(exp.amount for exp in deductible_expenses)
        
        # Calculate taxable income
        gross_income = revenue_data['total_revenue']
        taxable_income = gross_income - total_deductible
        
        # Group expenses by category for tax purposes
        expense_categories = {}
        for expense in deductible_expenses:
            if expense.category not in expense_categories:
                expense_categories[expense.category] = []
            expense_categories[expense.category].append({
                'date': expense.date.isoformat(),
                'description': expense.description,
                'amount': expense.amount
            })
        
        return {
            "period": {
                "start_date": start_date.strftime('%Y-%m-%d'),
                "end_date": end_date.strftime('%Y-%m-%d')
            },
            "income": {
                "gross_income": gross_income,
                "total_sales": revenue_data['total_sales']
            },
            "deductions": {
                "total_deductible_expenses": total_deductible,
                "expense_categories": expense_categories
            },
            "taxable_income": taxable_income,
            "estimated_tax_rate": 0.25,  # 25% estimated tax rate
            "estimated_tax_owed": taxable_income * 0.25
        }

    def get_recurring_expenses(self) -> List[Dict[str, Any]]:
        """Get all recurring expenses"""
        
        recurring = self.db.query(Expense).filter(
            Expense.is_recurring == True
        ).order_by(Expense.category, Expense.description).all()
        
        return [
            {
                "id": expense.id,
                "category": expense.category,
                "description": expense.description,
                "amount": expense.amount,
                "last_occurrence": expense.date.isoformat(),
                "frequency": "monthly"  # Default frequency
            }
            for expense in recurring
        ]

    def get_financial_insights(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """Get AI-powered financial insights"""
        
        financial_data = self.get_financial_summary(start_date, end_date)
        insights = []
        
        # Profit margin analysis
        profit_margin = financial_data['summary']['profit_margin']
        if profit_margin < 10:
            insights.append({
                "type": "warning",
                "title": "Low Profit Margin",
                "description": f"Your profit margin is {profit_margin:.1f}%, which is below the recommended 15-20%",
                "recommendation": "Consider reducing expenses or increasing prices"
            })
        elif profit_margin > 30:
            insights.append({
                "type": "positive",
                "title": "Healthy Profit Margin",
                "description": f"Your profit margin of {profit_margin:.1f}% is excellent",
                "recommendation": "Consider reinvesting profits into business growth"
            })
        
        # Expense analysis
        expense_breakdown = financial_data['expense_breakdown']
        if expense_breakdown:
            top_expense = expense_breakdown[0]
            if top_expense['percentage'] > 40:
                insights.append({
                    "type": "info",
                    "title": "High Expense Category",
                    "description": f"{top_expense['category']} represents {top_expense['percentage']:.1f}% of total expenses",
                    "recommendation": "Review this category for potential cost savings"
                })
        
        # Cash flow analysis
        cash_flow = self.get_cash_flow_analysis(start_date, end_date)
        net_cash_flow = cash_flow['summary']['net_cash_flow']
        
        if net_cash_flow < 0:
            insights.append({
                "type": "warning",
                "title": "Negative Cash Flow",
                "description": f"Your cash flow is negative by ${abs(net_cash_flow):.2f}",
                "recommendation": "Focus on increasing revenue or reducing expenses"
            })
        
        return insights